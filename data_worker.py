from openpyxl import load_workbook, workbook
import sqlite3
import fns
import json
import datetime
import time
import os
import re

api = None


def read_data(filename, logs):
    connection = sqlite3.connect("fns.db")
    cursor = connection.cursor()
    cursor.execute("DELETE FROM receipts")
    cursor.execute("DELETE FROM receipt_detail")
    wb = load_workbook(filename)
    sheet = wb.active
    max_row = sheet.max_row + 1
    for i in range(2, max_row):
        NN = i - 1
        if sheet.cell(row=i, column=1).value == None:
            continue
        summ = int(sheet.cell(row=i, column=1).value * 100)
        date_str = str(sheet.cell(row=i, column=2).value)
        date_str = date_str[6:10]+'-'+date_str[3:5]+'-'+date_str[0:2]
        time_str = str(sheet.cell(row=i, column=3).value)
        if time_str.count(":") == 1:
            time_str += ":00"
        time_list = time_str.split(':')
        for ind in range(len(time_list)):
            if len(time_list[ind].strip()) < 2:
                time_list[ind] = '0'+time_list[ind].strip()
        time_str = ":".join(time_list)
        date_time = (date_str+'T'+time_str).strip()
        if re.match(r'^[0-9]{4}\-[0-9]{2}\-[0-9]{2}T[0-9]{2}\:[0-9]{2}\:[0-9]{2}$', date_time) is None:
            logs.emit(f'Ошибка в формате даты/времени в строке {NN} date_time  = {date_time} ')
        fiscal_number = str(sheet.cell(row=i, column=4).value).replace('*', '')
        fiscal_data = str(sheet.cell(row=i, column=5).value).replace('*', '')
        fiscal_sign = str(sheet.cell(row=i, column=6).value).replace('*', '')
        txt_query = f"INSERT INTO receipts VALUES( '{NN}', '{summ}', '{date_time}', '{fiscal_number}', '{fiscal_data}','{fiscal_sign}', '','','','')"
        cursor.execute(txt_query)
        connection.commit()
    wb.close()
    cursor.execute("SELECT * from `receipts`")
    rows = cursor.fetchall()
    cursor.close()
    connection.close()
    return rows

def write_data(filename):
    error = None
    book = workbook.Workbook()
    sheet = book.active
    head = ['NN', 'Дата время', 'ФД', 'ФН', 'ФП', 'Сумма', 'ККТ', 'Адрес', 'ВремяИзЧека', 'Товар', 'Цена', 'Кол-во', 'Результат проверки']
    for j, col in enumerate(head):
        sheet.cell(row=1, column=j + 1).value = col
    connection = sqlite3.connect("fns.db")
    cursor = connection.cursor()
    sql_text = """
        SELECT  
        receipts.NN
        ,receipts.date_time
        ,receipts.fiscal_data
        ,receipts.fiscal_number
        ,receipts.fiscal_sign
        ,receipts.summ / 100.0
        ,receipts.kkt
        ,receipts.address
        ,receipts.dt_check
        ,receipt_detail.good
        ,receipt_detail.price
        ,receipt_detail.quantity
        ,receipts.result
    FROM receipts 
    LEFT JOIN receipt_detail ON receipts.NN = receipt_detail.receipt
    """
    cursor.execute(sql_text)
    rows = cursor.fetchall()
    i = 2
    for row in rows:
        for j, col in enumerate(row):
            sheet.cell(row=i, column=j+1).number_format = '@'
            sheet.cell(row=i, column=j+1).value = str(col).replace('None','')
        i += 1
        try:
            book.save(filename)
        except PermissionError:
            error = 'ошибка сохранения файла.закройте файл'
    cursor.close()
    connection.close()
    return error


def get_parameter(name_p):
    connection = sqlite3.connect('fns.db')
    cursor = connection.cursor()
    cursor.execute(f"SELECT `value` from `config` where `name`='{name_p}'")
    res = cursor.fetchone()
    value = res[0]
    cursor.close()
    connection.close()
    return value

def save_parameter(name_p,val_p):
    connection = sqlite3.connect('fns.db')
    cursor = connection.cursor()
    cursor.execute(f"UPDATE `config` set `value`='{val_p}' where `name`='{name_p}'")
    connection.commit()
    cursor.close()
    connection.close()

def collect_receipts(logs, delta_m =0):
    global api
    connection = sqlite3.connect("fns.db")
    cursor = connection.cursor()
    cursor.execute("SELECT * FROM receipts where `result`<>'чек корректный'")
    rows = cursor.fetchall()
    if len(rows) == 0:
        logs.emit("Нет чеков для проверки")
        return False
    for row in rows:
        message = {
            'status': 'error',
            'message':  '',
            'code': None
        }
        new_time = datetime.datetime.strptime(row[2], "%Y-%m-%dT%H:%M:%S")
        new_time += datetime.timedelta(minutes=delta_m)
        new_time = new_time.strftime("%Y-%m-%dT%H:%M:%S")
        NN = row[0]
        try:
            message = api.get_ticket(
                row[1],  # сумма чека в формате РРРКК, 12 рублей 23 копейки передавайте как 1223
                new_time,  # дата и время в формате %Y-%m-%dT%H:%M:%S
                row[3],  # код ККТ
                '1',  # тип операции
                row[4],  # номер фискального документа
                row[5]  # фискальный признак
            )
        except Exception as err_api:
            logs.emit(f"ошибка получения данных в фнс {err_api}")
            cursor.execute(f"UPDATE receipts set `result`='ошибка получения данных в фнс {err_api}' where `NN`={NN}")
            connection.commit()
        if message['code'] == '200':
            receipt = json.loads(message['message'])
            address = receipt['address'].replace("'", ' ')
            kkt = receipt['content']['kktRegId']
            dt = time.strftime("%H:%M", time.gmtime(receipt['content']['dateTime']))
            cursor.execute(f"UPDATE receipts set `dt_check`='{dt}', `kkt`='{kkt}',`address`='{address}', `result`='чек корректный' where `NN`={NN}")
            connection.commit()
            goods = receipt['content']['items']
            for good in goods:
                name = good['name'].replace("'", ' ')
                price = good['price']/100.0
                quantity = good['quantity']
                cursor.execute(f"INSERT INTO receipt_detail VALUES({NN},'{name}','{quantity}','{price}')")
                connection.commit()
            logs.emit(f"чек {NN} корректный")
        else:
            logs.emit(f"чек {NN} не найден")
            logs.emit(message['message'])
            cursor.execute(f"UPDATE receipts set `result`='{message['message']}' where `NN`='{NN}'")
            connection.commit()
    cursor.close()
    connection.close()
    api.set_counter()
    return True


def process1(logs):
    global api
    api = fns.FNSApi()
    message1 = api.get_session_token()
    if message1['status'] != 'success':
        logs.emit(f"Ошибка получения временного токена ' + message1['message']")
        return False
    if os.path.exists("fns.db") == False:
        logs.mysignal.emit("Отсутствует файл с базой fns.db")
        return False
    connection = sqlite3.connect("fns.db")
    cursor = connection.cursor()
    cursor.execute("DELETE FROM receipt_detail")
    cursor.close()
    connection.close()
    # проверяем чеки со временем как есть delta minute = 0
    collect_receipts(logs)


def process2(logs):
    global api
    if api.session_token is None:
        return False
    # проверка чеков с перебором минут
    logs.emit(f"Запускаем проверку чеков с перебором минут")
    delta_m = [-3, -2, -1, 1, 2, 3]
    for m in delta_m:
        logs.emit(f"Дельта минут {m}")
        if collect_receipts(logs, m) is False:
            break




